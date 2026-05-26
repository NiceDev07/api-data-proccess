"""
Tests para el endpoint POST /v2/first-rows y la lógica interna de preview.

Cubre:
- Lectura de CSV (con y sin encabezados, distintos delimitadores)
- Lectura de XLSX (con y sin encabezados)
- NullReader (file=None) y extensión no soportada
- Límite de filas (_MAX_ROWS=6) aplicado vía n_rows en el reader
- Seguridad: archivo no encontrado y path traversal
- Integración HTTP: respuesta con campo "data" (formato legacy frontend)
- Carga (stress): CSV de 100k filas y XLSX de 5k filas

Formato de respuesta:
    {
        "success": true,
        "data": [
            "col1;col2;col3",        ← encabezados unidos por delimitador
            "val1;val2;val3",        ← fila 1
            "val4;val5;val6",        ← fila 2
            ...
        ]
    }
    data[0]  = encabezados de columnas
    data[1:] = filas de datos (máx. _MAX_ROWS)

Parámetros configurables marcados con # <-- CAMBIAR PARA FORZAR FALLO
"""
from pathlib import Path

import polars as pl
import pytest
from fastapi import FastAPI
from httpx import AsyncClient, ASGITransport
from openpyxl import Workbook

from modules.process.app.files.factory import ReaderFileFactory
from modules.process.app.files.preview import _MAX_ROWS, get_first_rows
from modules.process.infrastructure.routes.preview import router as preview_router


# ── app mínima para tests de integración HTTP ─────────────────────────────────

def _make_app(base_dir: str) -> FastAPI:
    """Crea una FastAPI mínima apuntando al directorio temporal del test."""
    import config.settings as _s
    _s.settings.repository_files_dir = base_dir
    app = FastAPI()
    app.include_router(preview_router, prefix="/v2")
    return app


# ── helpers ───────────────────────────────────────────────────────────────────

def _write_bytes(tmp_path: Path, name: str, content: bytes) -> Path:
    """Escribe bytes en un archivo temporal y devuelve su Path."""
    p = tmp_path / name
    p.write_bytes(content)
    return p


def _write_xlsx(tmp_path: Path, name: str, rows: list[list], n_data_rows: int = 0) -> Path:
    """
    Crea un XLSX con openpyxl.
    rows: lista de listas — la primera es el encabezado.
    Si n_data_rows > 0, agrega esa cantidad de filas extra de datos.
    """
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    if n_data_rows:
        n_cols = len(rows[0]) if rows else 3
        for i in range(n_data_rows):
            ws.append([f"val_{i}_{c}" for c in range(n_cols)])
    p = tmp_path / name
    wb.save(str(p))
    return p


# ══════════════════════════════════════════════════════════════════════════════
# TestPreviewCSV — tests unitarios sobre get_first_rows con archivos CSV
# ══════════════════════════════════════════════════════════════════════════════

class TestPreviewCSV:

    @pytest.mark.anyio
    async def test_csv_headers_and_rows(self, tmp_path):
        """CSV normal con encabezados: data[0] contiene headers, data[1:] contiene filas."""
        _write_bytes(tmp_path, "data.csv", b"nombre;edad\nAna;30\nBob;25")
        result = await get_first_rows(str(tmp_path), "data.csv", ";", True, str(tmp_path))

        assert result["success"] is True
        # data[0] = encabezados unidos por el delimitador
        assert result["data"][0] == "nombre;edad"  # <-- CAMBIAR "nombre;edad" a "x;edad" para forzar fallo
        # data[1:] = filas de datos
        assert len(result["data"]) == 3             # 1 header + 2 filas
        assert result["data"][1] == "Ana;30"

    @pytest.mark.anyio
    async def test_csv_no_headers(self, tmp_path):
        """useHeaders=False: los headers se nombran column_1, column_2..."""
        _write_bytes(tmp_path, "nohead.csv", b"Ana;30\nBob;25")
        result = await get_first_rows(str(tmp_path), "nohead.csv", ";", False, str(tmp_path))

        assert result["success"] is True
        # Sin encabezado Polars asigna nombres automáticos; normalize_columns los convierte a column_N
        header_cols = result["data"][0].split(";")
        assert header_cols[0] == "column_1"  # <-- CAMBIAR a "column_0" para forzar fallo
        assert header_cols[1] == "column_2"

    @pytest.mark.anyio
    async def test_csv_partial_headers(self, tmp_path):
        """Primer encabezado vacío: la columna vacía pasa a column_1 por normalize_columns."""
        # Primer campo vacío — Polars lo deja como cadena vacía
        _write_bytes(tmp_path, "partial.csv", b";nombre;codigo\n300;Ana;A01")
        result = await get_first_rows(str(tmp_path), "partial.csv", ";", True, str(tmp_path))

        assert result["success"] is True
        # La primera columna tenía encabezado vacío → debe quedar como column_1
        header_cols = result["data"][0].split(";")
        assert header_cols[0] == "column_1"  # <-- CAMBIAR a "" para forzar fallo
        assert header_cols[1] == "nombre"

    @pytest.mark.anyio
    async def test_csv_max_rows(self, tmp_path):
        """CSV con 20 filas de datos: el preview devuelve solo _MAX_ROWS=6 filas de datos."""
        lines = ["col"] + [f"val{i}" for i in range(20)]
        _write_bytes(tmp_path, "big.csv", "\n".join(lines).encode())
        result = await get_first_rows(str(tmp_path), "big.csv", ";", True, str(tmp_path))

        # data[0] = header, data[1:] = filas (máx _MAX_ROWS)
        assert len(result["data"]) == _MAX_ROWS + 1  # <-- CAMBIAR _MAX_ROWS + 1 a _MAX_ROWS + 2 para forzar fallo

    @pytest.mark.anyio
    async def test_csv_all_cols_returned(self, tmp_path):
        """CSV con 15 columnas: el preview devuelve todas sin truncar."""
        headers = ";".join(f"col{i}" for i in range(15))
        values  = ";".join(str(i) for i in range(15))
        _write_bytes(tmp_path, "wide.csv", f"{headers}\n{values}".encode())
        result = await get_first_rows(str(tmp_path), "wide.csv", ";", True, str(tmp_path))

        # El backend devuelve todas las columnas — el frontend decide cuántas mostrar
        # data[0] son los 15 headers unidos por ";": 14 separadores
        assert len(result["data"][0].split(";")) == 15  # <-- CAMBIAR a 10 para forzar fallo

    @pytest.mark.anyio
    async def test_csv_semicolon_delimiter(self, tmp_path):
        """Delimitador punto y coma: las columnas se separan y unen correctamente."""
        _write_bytes(tmp_path, "semi.csv", b"a;b;c\n1;2;3")
        result = await get_first_rows(str(tmp_path), "semi.csv", ";", True, str(tmp_path))

        assert result["data"][0] == "a;b;c"
        assert result["data"][1] == "1;2;3"

    @pytest.mark.anyio
    async def test_csv_comma_delimiter(self, tmp_path):
        """Delimitador coma: las columnas se separan y unen correctamente con coma."""
        _write_bytes(tmp_path, "comma.csv", b"x,y,z\n7,8,9")
        result = await get_first_rows(str(tmp_path), "comma.csv", ",", True, str(tmp_path))

        assert result["data"][0] == "x,y,z"
        assert result["data"][1] == "7,8,9"


# ══════════════════════════════════════════════════════════════════════════════
# TestPreviewXLSX — tests unitarios sobre get_first_rows con archivos XLSX
# ══════════════════════════════════════════════════════════════════════════════

class TestPreviewXLSX:

    @pytest.mark.anyio
    async def test_xlsx_headers_and_rows(self, tmp_path):
        """XLSX normal: data[0] contiene headers, data[1:] contiene filas."""
        _write_xlsx(tmp_path, "data.xlsx", [
            ["nombre", "edad"],
            ["Ana", 30],
            ["Bob", 25],
        ])
        result = await get_first_rows(str(tmp_path), "data.xlsx", "", True, str(tmp_path))

        assert result["success"] is True
        # XLSX sin delimitador → se usa ";" por defecto para unir columnas
        assert result["data"][0] == "nombre;edad"  # <-- CAMBIAR "nombre;edad" a "x;edad" para forzar fallo
        assert len(result["data"]) == 3             # 1 header + 2 filas
        assert result["data"][1].split(";")[0] == "Ana"

    @pytest.mark.anyio
    async def test_xlsx_partial_headers(self, tmp_path):
        """Encabezado con primera celda vacía: column_1 por posición (usando openpyxl)."""
        wb = Workbook()
        ws = wb.active
        # Primera celda vacía — openpyxl la escribe como None
        ws.append([None, "nombre", "codigo"])
        ws.append([300000000, "Ana", "A01"])
        wb.save(str(tmp_path / "partial.xlsx"))

        result = await get_first_rows(str(tmp_path), "partial.xlsx", "", True, str(tmp_path))

        # La celda vacía es __UNNAMED__0 en Polars → normalize_columns lo convierte a column_1
        header_cols = result["data"][0].split(";")
        assert header_cols[0] == "column_1"  # <-- CAMBIAR a "" para forzar fallo
        assert header_cols[1] == "nombre"

    @pytest.mark.anyio
    async def test_xlsx_max_rows(self, tmp_path):
        """XLSX con 20 filas: el preview devuelve solo _MAX_ROWS=6 filas de datos."""
        _write_xlsx(tmp_path, "big.xlsx", [["col"]], n_data_rows=20)
        result = await get_first_rows(str(tmp_path), "big.xlsx", "", True, str(tmp_path))

        # data[0] = header, data[1:] = filas (máx _MAX_ROWS)
        assert len(result["data"]) == _MAX_ROWS + 1  # <-- CAMBIAR _MAX_ROWS + 1 a _MAX_ROWS + 2 para forzar fallo


# ══════════════════════════════════════════════════════════════════════════════
# TestPreviewNull — NullReader y extensión no soportada
# ══════════════════════════════════════════════════════════════════════════════

class TestPreviewNull:

    @pytest.mark.anyio
    async def test_null_reader_returns_empty(self):
        """NullReader (file=None en la factory) devuelve DataFrame vacío → headers=[], rows=[]."""
        factory = ReaderFileFactory()
        reader  = factory.create(None)  # file=None → NullReader

        # Construimos una config mínima (la ruta no importa para NullReader)
        from modules.process.domain.models.process_dto import BaseFileConfig
        config = BaseFileConfig(
            folder="/tmp",
            file="dummy.csv",
            delimiter=";",
            useHeaders=True,
            nameColumnDemographic="_",
        )
        lf = await reader.read(config)
        df = lf.collect().fill_null("")

        # DataFrame vacío → sin columnas ni filas
        assert df.columns == []   # <-- CAMBIAR a ["x"] para forzar fallo
        assert df.rows()   == []

    def test_unsupported_extension_raises(self):
        """Extensión .txt no soportada: ReaderFileFactory lanza FileNotFoundError."""
        factory = ReaderFileFactory()
        with pytest.raises(FileNotFoundError, match="FILE_NOT_FOUND"):
            factory.create("archivo.txt")  # <-- CAMBIAR a "archivo.csv" para forzar fallo


# ══════════════════════════════════════════════════════════════════════════════
# TestPreviewSecurity — path traversal y archivo inexistente
# ══════════════════════════════════════════════════════════════════════════════

class TestPreviewSecurity:

    @pytest.mark.anyio
    async def test_file_not_found_raises(self, tmp_path):
        """Archivo que no existe: lanza FileNotFoundError con código FILE_NOT_FOUND."""
        with pytest.raises(FileNotFoundError, match="FILE_NOT_FOUND"):
            await get_first_rows(str(tmp_path), "noexiste.csv", ";", True, str(tmp_path))

    @pytest.mark.anyio
    async def test_path_traversal_blocked(self, tmp_path):
        """Intento de path traversal con '../': lanza PermissionError con código ACCESS_DENIED."""
        with pytest.raises(PermissionError, match="ACCESS_DENIED"):
            # Intentamos salir del directorio autorizado usando '../'
            await get_first_rows(str(tmp_path), "../../etc/hosts", ";", True, str(tmp_path))


# ══════════════════════════════════════════════════════════════════════════════
# TestPreviewEndpoint — integración HTTP contra el router FastAPI
# ══════════════════════════════════════════════════════════════════════════════

class TestPreviewEndpoint:

    @pytest.mark.anyio
    async def test_endpoint_csv_ok(self, tmp_path):
        """CSV válido: respuesta 200 con campo 'data' (formato legacy frontend)."""
        _write_bytes(tmp_path, "test.csv", b"nombre;edad\nAna;30\nBob;25")
        app = _make_app(str(tmp_path))
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            resp = await client.post(
                "/v2/first-rows",
                json={"folder": str(tmp_path), "file": "test.csv", "delimiter": ";"},
            )

        assert resp.status_code == 200
        body = resp.json()
        assert body["success"] is True
        assert "data" in body                       # <-- CAMBIAR a "headers" para forzar fallo
        assert isinstance(body["data"], list)
        assert body["data"][0] == "nombre;edad"     # header como primer elemento

    @pytest.mark.anyio
    async def test_endpoint_xlsx_ok(self, tmp_path):
        """XLSX válido: respuesta 200 con campo 'data'."""
        _write_xlsx(tmp_path, "test.xlsx", [["col1", "col2"], ["val1", "val2"]])
        app = _make_app(str(tmp_path))
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            resp = await client.post(
                "/v2/first-rows",
                json={"folder": str(tmp_path), "file": "test.xlsx"},
            )

        assert resp.status_code == 200
        body = resp.json()
        assert body["success"] is True
        assert "data" in body
        assert isinstance(body["data"], list)

    @pytest.mark.anyio
    async def test_endpoint_invalid_extension(self, tmp_path):
        """Extensión .txt no soportada: respuesta 422 con mensaje INVALID_EXTENSION."""
        app = _make_app(str(tmp_path))
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            resp = await client.post(
                "/v2/first-rows",
                json={"folder": str(tmp_path), "file": "script.txt"},  # <-- CAMBIAR a "script.csv" para forzar fallo
            )

        assert resp.status_code == 422
        assert "INVALID_EXTENSION" in str(resp.json())

    @pytest.mark.anyio
    async def test_endpoint_empty_folder(self, tmp_path):
        """folder vacío: respuesta 422 con mensaje FIELD_REQUIRED."""
        app = _make_app(str(tmp_path))
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            resp = await client.post(
                "/v2/first-rows",
                json={"folder": "", "file": "test.csv"},  # <-- CAMBIAR "" a un path válido para forzar fallo
            )

        assert resp.status_code == 422
        assert "FIELD_REQUIRED" in str(resp.json())

    @pytest.mark.anyio
    async def test_endpoint_file_not_found(self, tmp_path):
        """Archivo que no existe en disco: respuesta 404 con FILE_NOT_FOUND."""
        app = _make_app(str(tmp_path))
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            resp = await client.post(
                "/v2/first-rows",
                json={"folder": str(tmp_path), "file": "noexiste.csv"},  # <-- CAMBIAR a un archivo que sí existe para forzar fallo
            )

        assert resp.status_code == 404
        assert "FILE_NOT_FOUND" in resp.json()["detail"]


# ══════════════════════════════════════════════════════════════════════════════
# TestPreviewBulk — tests de carga (stress)
# ══════════════════════════════════════════════════════════════════════════════

class TestPreviewBulk:

    @pytest.mark.anyio
    async def test_bulk_csv_only_returns_max_rows(self, tmp_path):
        """CSV de 100 000 filas: el preview devuelve solo _MAX_ROWS=6 filas de datos."""
        # Generamos un CSV grande en disco
        N_ROWS = 100_000  # <-- CAMBIAR a 5 para forzar fallo (daría menos de _MAX_ROWS)
        csv_path = tmp_path / "bulk.csv"
        with open(csv_path, "w") as f:
            f.write("telefono;nombre;codigo\n")
            for i in range(N_ROWS):
                f.write(f"{300000000 + i};nombre_{i};cod_{i}\n")

        result = await get_first_rows(str(tmp_path), "bulk.csv", ";", True, str(tmp_path))

        # Sin importar cuántas filas tenga el CSV, la preview no supera _MAX_ROWS datos
        assert result["success"] is True
        # data[0] = header, data[1:] = filas → total = _MAX_ROWS + 1
        assert len(result["data"]) == _MAX_ROWS + 1  # <-- CAMBIAR _MAX_ROWS + 1 a _MAX_ROWS + 2 para forzar fallo

    @pytest.mark.anyio
    async def test_bulk_xlsx_only_returns_max_rows(self, tmp_path):
        """XLSX de 5 000 filas: el preview devuelve solo _MAX_ROWS=6 filas de datos."""
        N_ROWS = 5_000  # <-- CAMBIAR a 3 para forzar fallo (daría menos de _MAX_ROWS)
        wb = Workbook()
        ws = wb.active
        ws.append(["telefono", "nombre", "codigo"])
        for i in range(N_ROWS):
            ws.append([300000000 + i, f"nombre_{i}", f"cod_{i}"])
        wb.save(str(tmp_path / "bulk.xlsx"))

        result = await get_first_rows(str(tmp_path), "bulk.xlsx", "", True, str(tmp_path))

        assert result["success"] is True
        # data[0] = header, data[1:] = filas → total = _MAX_ROWS + 1
        assert len(result["data"]) == _MAX_ROWS + 1  # <-- CAMBIAR _MAX_ROWS + 1 a _MAX_ROWS + 2 para forzar fallo
